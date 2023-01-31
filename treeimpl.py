import openpyxl

path="zerprodf.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
maxrow=sheet_obj.max_row
maxcolumn=sheet_obj.max_column
class Head:
    def __init__(self):
        self.next=[]
class Normal:
    def __init__(self,name,value):
        self.name=name
        self.value=value
        self.next=[]

ageBucket=Head()
temp_list=[]
for i in range(2,maxrow+1):
    name=sheet_obj.cell(row=i,column=1).value
    value=sheet_obj.cell(row=i,column=2).value
    agenode=Normal(name,value)
    temp_list.append(agenode)
ageBucket.next=temp_list
features=[2,2,2,2,3,2,2,2]
index_done=-1


queue_current=[]
child_current=[]
temp=[]
for i in ageBucket.next:
    queue_current.append(i)

def insert_node(node, index_done):
    columns_done = 2
    for i in range(index_done + 1):
        columns_done += features[i]
    temp_list = []
    for i in range(features[index_done + 1]):
        name = sheet_obj.cell(row=1, column=columns_done + i + 1).value
        value = sheet_obj.cell(row=2, column=columns_done + i + 1).value
        newnode = Normal(name, value)
        temp_list.append(newnode)
        global child_current
        child_current.append(newnode)
    node.next = temp_list
    #print(f"for {node.name} children are:")
    # for i in node.next:
    #     print(i.name, end=" ")
    # print("\n")
    return

def addchildren(thelist,index_done):
    if index_done==len(features)-1:
        return
    for i in thelist:
        insert_node(i,index_done)
    global queue_current,child_current
    queue_current=child_current
    child_current=[]
    #index_done+=1
    return addchildren(queue_current,index_done+1)

addchildren(queue_current,-1)

#<-------------------------------------Following is for tree traversal through all possible paths------------------------------------------------->
val=0
total=0
scoredict={}
def treetraverse(node,pathvalue,path_list):
    global val,total,scoredict
    # print("[",end="")
    # for i in node:
    #     print(i.name,end=",")
    # print("]")
    if node==[]:
        #print(round(pathvalue,3))
        path_string=""
        for i in path_list:
                path_string+=i+"->"
        scoredict[path_string]=[round(pathvalue,3)]
        #print(path_list)
        #print("=",round(pathvalue,3))
        total+=1
        if round(pathvalue,3)>val:
            val=round(pathvalue,3)
        else:
            pass
        return
    # if node[0].next==[]:
    #     print(round(pathvalue,3))
    #     total+=1
    #     if round(pathvalue,3)>val:
    #         val=round(pathvalue,3)
    #     else:
    #         pass
    #     return
    temp=path_list.copy()
    temp.append(node[0].name)
    treetraverse(node[0].next,pathvalue+node[0].value,temp)
    if len(node)==1:
        return
    treetraverse(node[1:],pathvalue,path_list)

treetraverse(ageBucket.next,0,[])
# for keys in scoredict:
#     print(keys)
#     print("=",scoredict[keys])
for keys in scoredict:
    print(keys)
print("maximum value is: ",val)
print("Total path taken: ",total)