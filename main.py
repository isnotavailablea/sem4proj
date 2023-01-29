import openpyxl
path="withang.xlsx"
wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
cell_obj=sheet_obj.cell(row=1,column=1)
# print(cell_obj.value)
totalrows=896
maindict = {"ageGroups": {
        "26-30": {"value": 0, "others": {}}, "31-35": {"value": 0, "others": {}}, "36-40": {"value": 0, "others": {}},
        "41-45": {"value": 0, "others": {}}, "46-50": {"value": 0, "others": {}}, "51-55": {"value": 0, "others": {}},
        "56-60": {"value": 0, "others": {}}, "61-65": {"value": 0, "others": {}}, "66-70": {"value": 0, "others": {}}
        , "71-75": {"value": 0, "others": {}}, "76-80": {"value": 0, "others": {}}, "81-85": {"value": 0, "others": {}},
        "86-90": {"value": 0, "others": {}}
    }}


def agecategory(age):
    if age<=25:
        return "26-30"
    if age>=90:
        return "86-90"
    for keys in maindict["ageGroups"]:
        l = keys.split("-")
       # print(l)
        if age >= int(l[0]) and age <= int(l[1]):
            return keys
    return "weird "+str(age)
# for i in range(2,totalrows+1):
#     if(sheet_obj.cell(row=i,column=2).value==1):
#         maindict["ageGroups"][agecategory(sheet_obj(row=i,column=1).value)]+=1
tot=0
for i in range(2, totalrows + 1):
    if(sheet_obj.cell(row=i,column=2).value==0):
        continue
    tot+=1
    # print(sheet_obj.cell(row=i,column=1).value)
    agecat=agecategory(sheet_obj.cell(row=i,column=1).value)
    maindict["ageGroups"][agecat]["value"]+=1
    for column in range(2,11):
            # print(i)
            if column==9:
                agecat=agecategory(sheet_obj.cell(row=i,column=1).value)
                try:
                    if sheet_obj.cell(row=i,column=9).value>=124 and sheet_obj.cell(row=i,column=9).value<=194:
                           if "chol0" not in maindict["ageGroups"][agecat]["others"]:
                               maindict["ageGroups"][agecat]["others"]["chol0"]=1
                           else:
                               maindict["ageGroups"][agecat]["others"]["chol0"]+=1
                    else:
                        if "chol1" in maindict["ageGroups"][agecat]["others"]:
                            maindict["ageGroups"][agecat]["others"]["chol1"]+=1
                        else:
                            maindict["ageGroups"][agecat]["others"]["chol1"]=1
                except:
                    print(agecat)
                continue
            if(column==5 or column==6):
                continue
            if((column==4)):
                agecat = agecategory(sheet_obj.cell(row=i, column=1).value)
                [a,b,c]=[sheet_obj.cell(row=i,column=5).value,sheet_obj.cell(row=i,column=6).value,sheet_obj.cell(row=i,column=7).value]
                if a>=0.7:
                    if "ang11" not in maindict["ageGroups"][agecat]["others"]:
                        maindict["ageGroups"][agecat]["others"]["ang11"]=1
                    else:
                        maindict["ageGroups"][agecat]["others"]["ang11"]+=1
                else:
                    if "ang10" not in maindict["ageGroups"][agecat]["others"]:
                        maindict["ageGroups"][agecat]["others"]["ang10"]=1
                    else:
                        maindict["ageGroups"][agecat]["others"]["ang10"]+=1
                if b>=0.7:
                    if "ang21" not in maindict["ageGroups"][agecat]["others"]:
                        maindict["ageGroups"][agecat]["others"]["ang21"]=1
                    else:
                        maindict["ageGroups"][agecat]["others"]["ang21"]+=1
                else:
                    if "ang20" not in maindict["ageGroups"][agecat]["others"]:
                        maindict["ageGroups"][agecat]["others"]["ang20"]=1
                    else:
                        maindict["ageGroups"][agecat]["others"]["ang20"]+=1
                if c>=0.7:
                    if "ang31" not in maindict["ageGroups"][agecat]["others"]:
                        maindict["ageGroups"][agecat]["others"]["ang31"]=1
                    else:
                        maindict["ageGroups"][agecat]["others"]["ang31"]+=1
                else:
                    if "ang30" not in maindict["ageGroups"][agecat]["others"]:
                        maindict["ageGroups"][agecat]["others"]["ang30"]=1
                    else:
                        maindict["ageGroups"][agecat]["others"]["ang30"]+=1
                continue
            agecat = agecategory(sheet_obj.cell(row=i, column=1).value)
            stdi=sheet_obj.cell(row=1, column=column).value + "-" + str(sheet_obj.cell(row=i, column=column).value)
           # print(i," ",column," ",stdi)
            try:
                if str(stdi) in maindict["ageGroups"][agecat]["others"]:
                    maindict["ageGroups"][agecat]["others"][sheet_obj.cell(row=1, column=column).value + "-" + str(
                        sheet_obj.cell(row=i, column=column).value)] += 1
                else:
                    maindict["ageGroups"][agecat]["others"][sheet_obj.cell(row=1, column=column).value + "-" + str(
                        sheet_obj.cell(row=i, column=column).value)] = 1
            except Exception as e:
                print(agecat)


a={"sample"}
for key in maindict["ageGroups"]:
    for k in maindict["ageGroups"][key]["others"]:
        a.add(k)
# a.sort()
# print(a)
# print(tot)
# print(maindict)
